import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# --- 读取文本文件内容 ---
try:
    with open('accuracy_results.txt', 'r', encoding='utf-8') as file:
        content = file.read()
except FileNotFoundError:
    print("错误：'accuracy_results.txt' 文件未找到。")
    exit()

# --- 数据提取和处理 ---
# 定义数据列名和对应的正则表达式
patterns = {
    'base_vector': r'Processing directory: base_vector.*?All re-evaluation tasks completed.',
    'gt_base_vector': r'Processing directory: gt_base_vector.*?All re-evaluation tasks completed.',
    'licv_version': r'Processing directory: licv_version.*?All re-evaluation tasks completed.',
    'gt_licv_version': r'Processing directory: gt_licv_version.*?All re-evaluation tasks completed.',
    'mimic_vector': r'Processing directory: mimic_vector.*?All re-evaluation tasks completed.',
    'gt_mimic_vector': r'Processing directory: gt_mimic_vector.*?All re-evaluation tasks completed.'
}

# 提取 '0-shot' 数据的正则表达式
zero_shot_pattern = r'Processing single file: icl-llama-3.1-8b-instruct-gsm8k-0shot-direct.*?Accuracy: ([\d.]+) %'

# 用于存储所有数据的字典
results = {
    'Self generate:base_version': {},
    'GT:base_version': {},
    'Self generate:licv_version': {},
    'GT:licv_version': {},
    'Self generate:mimic_version': {},
    'GT:mimic_version': {}
}

# 遍历每个版本，提取其所有 layer 的数据
for name, pattern in patterns.items():
    # 找到该版本对应的文本块
    block_match = re.search(pattern, content, re.DOTALL)
    if not block_match:
        continue
    block_content = block_match.group(0)

    # 提取所有 layer 的准确率
    layer_matches = re.findall(r'layer (\d+).*?Accuracy: ([\d.]+)', block_content)
    
    for layer, accuracy in layer_matches:
        layer_name = f'layer{layer}'
        accuracy_value = float(accuracy)

        # 根据 name 映射到正确的列名
        if name == 'base_vector':
            results['Self generate:base_version'][layer_name] = accuracy_value
        elif name == 'gt_base_vector':
            results['GT:base_version'][layer_name] = accuracy_value
        elif name == 'licv_version':
            results['Self generate:licv_version'][layer_name] = accuracy_value
        elif name == 'gt_licv_version':
            results['GT:licv_version'][layer_name] = accuracy_value
        elif name == 'mimic_vector':
            results['Self generate:mimic_version'][layer_name] = accuracy_value
        elif name == 'gt_mimic_vector':
            results['GT:mimic_version'][layer_name] = accuracy_value

# 将数据转换成 DataFrame
df = pd.DataFrame(results)

# 确保索引（即 layer）是按数字顺序排列的
df.index = pd.Index([int(re.search(r'\d+', idx).group(0)) for idx in df.index], name='Layer')
df = df.sort_index()
df.index = [f'layer{i}' for i in df.index]

# 处理 '0-shot' 数据
zero_shot_match = re.search(zero_shot_pattern, content, re.DOTALL)
if zero_shot_match:
    zero_shot_accuracy = float(zero_shot_match.group(1))
    df.loc['0shot'] = [zero_shot_accuracy, zero_shot_accuracy, zero_shot_accuracy, zero_shot_accuracy, zero_shot_accuracy, zero_shot_accuracy]

# 将 '0-shot' 放到最后一行
if '0shot' in df.index:
    df = pd.concat([df.drop('0shot'), df.loc[['0shot']]])

# --- Excel生成部分 ---
wb = Workbook()
ws = wb.active
ws.title = "Accuracy Results"

headers = ["Layer", "Self generate:base_version", "GT:base_version", 
           "Self generate:licv_version", "GT:licv_version",
           "Self generate:mimic_version", "GT:mimic_version"]
# 检查 '0shot' 是否需要额外添加
if '0shot' in df.columns:
    headers.append('0-shot')

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 写入数据
for i, layer in enumerate(df.index):
    ws.cell(row=i + 2, column=1, value=layer)
    for j, col_name in enumerate(df.columns, 2):
        ws.cell(row=i + 2, column=j, value=df.loc[layer, col_name])

# 调整列宽
for col in range(1, len(headers) + 1):
    ws.column_dimensions[chr(64 + col)].width = 20

wb.save("accuracy_results.xlsx")
print("Excel is over:accuracy_results.xlsx")